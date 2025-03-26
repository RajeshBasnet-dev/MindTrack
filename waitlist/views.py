from django.shortcuts import render
from django.http import JsonResponse
from .forms import WaitlistForm
from .models import WaitlistEntry, CommunityPost, Reply
from openpyxl import Workbook, load_workbook
import os
from django.conf import settings

def home(request):
    if request.method == 'POST':
        form = WaitlistForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            if WaitlistEntry.objects.filter(email=email).exists():
                return JsonResponse({'error': 'This email is already registered!'}, status=200)
            entry = form.save()
            excel_file = os.path.join(settings.BASE_DIR, 'waitlist.xlsx')
            if not os.path.exists(excel_file):
                wb = Workbook()
                ws = wb.active
                ws.append(['Name', 'Email', 'Preferences'])
            else:
                wb = load_workbook(excel_file)
                ws = wb.active
            ws.append([
                form.cleaned_data['name'],
                form.cleaned_data['email'],
                form.cleaned_data['preferences']
            ])
            wb.save(excel_file)
            return JsonResponse({'message': "Thanks for signing up! We'll notify you when the website is ready!"}, status=200)
        else:
            if 'email' in form.errors and any("already exists" in error for error in form.errors['email']):
                return JsonResponse({'error': 'This email is already registered!'}, status=200)
            return JsonResponse({'error': 'Invalid form data', 'details': form.errors.as_json()}, status=200)
    else:
        form = WaitlistForm()
    return render(request, 'waitlist/home.html', {'form': form})

def learnmore(request):
    return render(request, 'waitlist/learnmore.html')

def community(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        name = request.POST.get('name', 'Anonymous')
        message = request.POST.get('message')
        post_id = request.POST.get('post_id')  # Check if this is a reply

        if not email or not message:
            return JsonResponse({'error': 'Email and message are required!'}, status=200)

        if post_id:  # Handle reply
            try:
                post = CommunityPost.objects.get(id=post_id)
                reply = Reply(post=post, email=email, name=name, message=message)
                reply.save()

                # Save reply to Excel
                excel_file = os.path.join(settings.BASE_DIR, 'community_posts.xlsx')
                if not os.path.exists(excel_file):
                    wb = Workbook()
                    ws = wb.active
                    ws.append(['Post ID', 'Email', 'Name', 'Message', 'Timestamp', 'Type'])
                else:
                    wb = load_workbook(excel_file)
                    ws = wb.active
                ws.append([post_id, email, name, message, reply.created_at.strftime('%Y-%m-%d %H:%M:%S'), 'Reply'])
                wb.save(excel_file)

                return JsonResponse({'message': 'Your reply has been added!'}, status=200)
            except CommunityPost.DoesNotExist:
                return JsonResponse({'error': 'Post not found!'}, status=200)
        else:  # Handle new post
            post = CommunityPost(email=email, name=name, message=message)
            post.save()

            excel_file = os.path.join(settings.BASE_DIR, 'community_posts.xlsx')
            if not os.path.exists(excel_file):
                wb = Workbook()
                ws = wb.active
                ws.append(['Post ID', 'Email', 'Name', 'Message', 'Timestamp', 'Type'])
            else:
                wb = load_workbook(excel_file)
                ws = wb.active
            ws.append([post.id, email, name, message, post.created_at.strftime('%Y-%m-%d %H:%M:%S'), 'Post'])
            wb.save(excel_file)

            return JsonResponse({'message': 'Your post has been added to the community!'}, status=200)

    posts = CommunityPost.objects.all().order_by('-created_at')[:10]  # Last 10 posts
    return render(request, 'waitlist/community.html', {'posts': posts})